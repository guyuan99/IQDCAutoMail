# 08/16/2017 Created by Yuan Gu

import ftplib
import os.path, time
import datetime as dt
import openpyxl

#Open ftp connection
print "\n************ Connecting to FTP... ****************************************************\n"
ftp = ftplib.FTP('10.18.232.21', 'iqdcas2','OS_cdqi4')
#Get to the dir
ftp.cwd("/inboundFromEFOX/INDY/bak")
#List the files in the current directory
print "\n************ Pulling all file names from FTP... **************************************\n"
ls=[]
files = ftp.nlst()
for file in files:
	if(file[-3:]=="TXT"):
		ls.append(file)

#print ls
ftp.quit()
print "\n************ Creating new excel for pasting updated data... **************************\n"
#Create another excel without macro 
os.system("copy template.xlsx " + "IQDCSummaryReport_noMacro.xlsx")

print "\n************ Running macro and copying updated data to the new workbook... *********** \n"
#Run excel macro get IQDC of last week and copy data here
os.system("start "+"IQDCSummaryReport.xlsm")

print "\n************ Updating FTP file name in Excel by comparing date and file name... ****** \n"
#Compare date with the name of the file from FTP and update specific cell in the workbook
index=7
today = dt.date.today()
lastWeekday=today - dt.timedelta(days=today.weekday())-dt.timedelta(days=index)
lastWeekday=lastWeekday.strftime('%y%m%d')

wb = openpyxl.load_workbook('IQDCSummaryReport_noMacro.xlsx')
worksheet = wb.get_sheet_by_name("Sheet1")
for i in range(7):
	nameMaker="INDY_20"+lastWeekday+"120600_IQDC.XVSFDMTYP.TXT"
	#print nameMaker
	if nameMaker in ls:
		worksheet.cell(row=i+2,column=4).value=nameMaker
	else:
		print "Did not find " + nameMaker+ "!\n"
		worksheet.cell(row=i+2,column=4).value="Information either in next TXT file or last TXT file."
	index-=1
	lastWeekday=today - dt.timedelta(days=today.weekday())-dt.timedelta(days=index)
	lastWeekday=lastWeekday.strftime('%y%m%d')

wb.save("backupIQDCfile/IQDCSummaryReport " + dt.datetime.today().strftime("%m_%d_%Y") + ".xlsx")
wb.save("IQDCSummaryReport_noMacro.xlsx")

print "\n************ Sending email...********************************************************* \n"

import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders

SERVER = "10.18.210.246"
FROM = "Yuan.Gu@foxconn.com"
TO = ["xxxx@abc.com","sse-it@abc.com","john.markman@abc.com","xxxx@abc.com","xxxx@abc.com"] # must be a list
#TO = ["Yuan.Gu@abc.com"]
TOStringMaker=""
for str in TO:
	TOStringMaker=TOStringMaker+str+";"

msg = MIMEMultipart()
 
msg['From'] = FROM
msg['To'] = TOStringMaker
msg['Subject'] = "FXINDY IQDC Weekly roll-up file " + dt.datetime.today().strftime("%Y-%m")
 
body = "Hi John,\n\nPlease find the attached weekly IQDC report. "
 
msg.attach(MIMEText(body, 'plain'))


filename = "IQDCSummaryReport " + dt.datetime.today().strftime("%m_%d_%Y") + ".xlsx"
attachment = open("backupIQDCfile\IQDCSummaryReport " + dt.datetime.today().strftime("%m_%d_%Y") + ".xlsx", "rb")
 
part = MIMEBase('application', 'octet-stream')
part.set_payload((attachment).read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
msg.attach(part)
message= msg.as_string()

# Send the mail
server = smtplib.SMTP(SERVER)
server.sendmail(FROM, TO, message)
server.quit()
